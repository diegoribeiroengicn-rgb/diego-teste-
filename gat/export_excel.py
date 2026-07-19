"""
Exportação de relatórios em Excel (.xlsx), respeitando a estrutura e o
layout das planilhas originais fornecidas pela Tecnoplano, divididas em
abas: Prestadores, Cessionários, Alertas Críticos (Pendente de Reunião) e
Painel Consolidado de Dados.
"""

from __future__ import annotations

import io

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from gat.business_rules import enriquecer_cessionarios, enriquecer_prestadores, filtrar_ativos
from gat.config import COLUNAS_EXIBICAO_CESSIONARIOS, COLUNAS_EXIBICAO_PRESTADORES
from gat.database import listar_cessionarios, listar_prestadores

_COR_CABECALHO = "1B3A8A"  # Navy Tecnoplano
_FONTE_CABECALHO = Font(name="Arial", bold=True, color="FFFFFF", size=11)
_FONTE_CORPO = Font(name="Arial", size=10)
_PREENCHIMENTO_CABECALHO = PatternFill(start_color=_COR_CABECALHO, end_color=_COR_CABECALHO, fill_type="solid")
_BORDA_FINA = Border(*(Side(style="thin", color="D9D9D9") for _ in range(4)))


def _escrever_dataframe(ws: Worksheet, df: pd.DataFrame, titulo: str) -> None:
    """Escreve um DataFrame em uma planilha, com cabeçalho estilizado e colunas ajustadas."""
    ws["A1"] = titulo
    ws["A1"].font = Font(name="Arial", bold=True, size=13, color=_COR_CABECALHO)
    linha_cabecalho = 3

    for col_idx, coluna in enumerate(df.columns, start=1):
        celula = ws.cell(row=linha_cabecalho, column=col_idx, value=str(coluna))
        celula.font = _FONTE_CABECALHO
        celula.fill = _PREENCHIMENTO_CABECALHO
        celula.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        celula.border = _BORDA_FINA

    for row_idx, linha in enumerate(df.itertuples(index=False), start=linha_cabecalho + 1):
        for col_idx, valor in enumerate(linha, start=1):
            celula = ws.cell(row=row_idx, column=col_idx, value=valor)
            celula.font = _FONTE_CORPO
            celula.border = _BORDA_FINA

    for col_idx, coluna in enumerate(df.columns, start=1):
        largura = max(12, min(40, len(str(coluna)) + 4))
        ws.column_dimensions[get_column_letter(col_idx)].width = largura

    ws.freeze_panes = ws.cell(row=linha_cabecalho + 1, column=1)
    if len(df) > 0:
        ultima_coluna = get_column_letter(len(df.columns))
        ws.auto_filter.ref = f"A{linha_cabecalho}:{ultima_coluna}{linha_cabecalho}"


def gerar_relatorio_excel() -> bytes:
    """
    Gera o relatório consolidado em Excel com as abas: PRESTADORES,
    CESSIONARIOS, ALERTAS CRÍTICOS e PAINEL CONSOLIDADO, prontas para
    download pelo usuário.
    """
    df_prest = enriquecer_prestadores(listar_prestadores())
    df_cess = enriquecer_cessionarios(listar_cessionarios())

    df_prest_ativos = filtrar_ativos(df_prest)
    df_cess_ativos = filtrar_ativos(df_cess)

    wb = Workbook()
    wb.remove(wb.active)

    # --- Aba PRESTADORES (dados completos, incl. cancelados, para fidelidade histórica) ---
    ws_prest = wb.create_sheet("PRESTADORES")
    colunas_prest = list(COLUNAS_EXIBICAO_PRESTADORES.keys())
    df_prest_export = df_prest[colunas_prest].rename(columns=COLUNAS_EXIBICAO_PRESTADORES) if not df_prest.empty else pd.DataFrame(columns=list(COLUNAS_EXIBICAO_PRESTADORES.values()))
    _escrever_dataframe(ws_prest, df_prest_export, "ANÁLISE DE PROJETOS - PRESTADORES DE SERVIÇO")

    # --- Aba CESSIONARIOS ---
    ws_cess = wb.create_sheet("CESSIONARIOS")
    colunas_cess = list(COLUNAS_EXIBICAO_CESSIONARIOS.keys())
    df_cess_export = df_cess[colunas_cess].rename(columns=COLUNAS_EXIBICAO_CESSIONARIOS) if not df_cess.empty else pd.DataFrame(columns=list(COLUNAS_EXIBICAO_CESSIONARIOS.values()))
    _escrever_dataframe(ws_cess, df_cess_export, "ANÁLISE DE PROJETOS - CESSIONÁRIOS")

    # --- Aba ALERTAS CRÍTICOS (Pendente de Reunião, apenas projetos ativos) ---
    ws_alertas = wb.create_sheet("ALERTAS CRITICOS")
    pendentes_prest = df_prest_ativos[df_prest_ativos["pendente_reuniao"]] if not df_prest_ativos.empty else df_prest_ativos
    pendentes_cess = df_cess_ativos[df_cess_ativos["pendente_reuniao"]] if not df_cess_ativos.empty else df_cess_ativos

    linhas_alerta = []
    for _, linha in pendentes_prest.iterrows():
        linhas_alerta.append({
            "Origem": "Prestador", "Item": linha.get("item"), "Nome": linha.get("prestador"),
            "Disciplina": linha.get("disciplina"), "Revisão": linha.get("revisao"),
            "Responsável": linha.get("responsavel"), "Status Análise": linha.get("status_analise"),
            "Observações": linha.get("observacoes"),
        })
    for _, linha in pendentes_cess.iterrows():
        linhas_alerta.append({
            "Origem": "Cessionário", "Item": linha.get("item"), "Nome": linha.get("cessionario"),
            "Disciplina": linha.get("disciplina"), "Revisão": linha.get("revisao"),
            "Responsável": linha.get("responsavel"), "Status Análise": linha.get("status_analise"),
            "Observações": linha.get("observacoes"),
        })
    df_alertas = pd.DataFrame(linhas_alerta) if linhas_alerta else pd.DataFrame(
        columns=["Origem", "Item", "Nome", "Disciplina", "Revisão", "Responsável", "Status Análise", "Observações"]
    )
    _escrever_dataframe(ws_alertas, df_alertas, "ALERTAS CRÍTICOS — PENDENTE DE REUNIÃO (REV2+)")

    # --- Aba PAINEL CONSOLIDADO ---
    ws_painel = wb.create_sheet("PAINEL CONSOLIDADO")
    total_prest = len(df_prest_ativos)
    total_cess = len(df_cess_ativos)
    atrasados_prest = int((df_prest_ativos["status_entrega_calc"] == "ATRASADO").sum()) if not df_prest_ativos.empty else 0
    atrasados_cess = int((df_cess_ativos["status_entrega_calc"] == "ATRASADO").sum()) if not df_cess_ativos.empty else 0
    pendentes_reuniao_total = len(pendentes_prest) + len(pendentes_cess)

    resumo = pd.DataFrame([
        {"Indicador": "Projetos Ativos - Prestadores", "Valor": total_prest},
        {"Indicador": "Projetos Ativos - Cessionários", "Valor": total_cess},
        {"Indicador": "Total de Projetos Ativos", "Valor": total_prest + total_cess},
        {"Indicador": "Atrasados - Prestadores", "Valor": atrasados_prest},
        {"Indicador": "Atrasados - Cessionários", "Valor": atrasados_cess},
        {"Indicador": "Pendente de Reunião (REV2+)", "Valor": pendentes_reuniao_total},
    ])
    _escrever_dataframe(ws_painel, resumo, "DASHBOARD GAT — PAINEL CONSOLIDADO")

    if not df_prest_ativos.empty:
        resumo_resp_prest = df_prest_ativos.groupby("responsavel").size().reset_index(name="Projetos (Prestadores)")
        inicio_linha = 6 + len(resumo) + 3
        ws_painel.cell(row=inicio_linha - 2, column=1, value="Projetos por Responsável — Prestadores").font = Font(bold=True, color=_COR_CABECALHO)
        for col_idx, coluna in enumerate(resumo_resp_prest.columns, start=1):
            ws_painel.cell(row=inicio_linha, column=col_idx, value=str(coluna)).font = Font(bold=True)
        for row_offset, linha in enumerate(resumo_resp_prest.itertuples(index=False), start=1):
            for col_idx, valor in enumerate(linha, start=1):
                ws_painel.cell(row=inicio_linha + row_offset, column=col_idx, value=valor)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
