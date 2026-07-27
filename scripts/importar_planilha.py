#!/usr/bin/env python3
"""
Importação dos dados históricos reais das planilhas oficiais da Tecnoplano
para o banco de dados SQLite do Sistema GAT 2026.

Lê as abas `PROJ_PREST` e `PROJ_CESS` do arquivo `Controle_GAT_Projetos_2026.xlsm`
e a aba `AVALIAÇÃO_PRESTADORES` do arquivo `Avaliacao_Prestadores_GAT.xlsx`,
mapeando exatamente as colunas identificadas na engenharia reversa, e grava
os registros no banco — para que a aplicação já nasça povoada com o
histórico real, restando ao usuário cadastrar apenas os projetos novos.

Uso:
    python scripts/importar_planilha.py --xlsm <caminho.xlsm> [--avaliacao <caminho.xlsx>] --destino <banco.db>

    Para atualizar os dados a partir de uma planilha mais recente (banco já
    povoado): adicione --atualizar — cria um backup de segurança do banco
    atual, apaga apenas os registros de prestadores/cessionarios (usuários,
    reuniões, avaliações de checklist, planos de ação e configurações não
    são afetados) e reimporta do zero a partir da planilha informada.

Regras aplicadas (decisões registradas conforme engenharia reversa):
* Linhas sem nome do prestador/cessionário são ignoradas (linhas em branco
  da planilha).
* Colunas calculadas dinamicamente pelo sistema (hold em dias, tempo de
  análise, status de entrega) NÃO são importadas como valor fixo — são
  recalculadas em tempo real pelo motor de regras (`gat.business_rules`),
  evitando divergência entre o valor congelado da fórmula do Excel e o
  valor real na data de uso do sistema.
* Os campos "Mês", "Ano" e "Mês da Demanda" também não são armazenados
  como colunas redundantes: são derivados das datas em tempo de exibição.
* Valores de status/discipline fora do vocabulário padrão são preservados
  como texto livre (não são descartados nem "corrigidos" silenciosamente).
* A importação é idempotente por padrão: se o banco de destino já possuir
  registros nas tabelas de projetos, o script recusa a execução para
  evitar duplicidade — utilize --forcar para sobrescrever mesmo assim.
"""

from __future__ import annotations

import argparse
import sqlite3
import sys
from datetime import date, datetime
from pathlib import Path

import openpyxl

USUARIO_IMPORTACAO = "IMPORTACAO_PLANILHA"


def _valor_data(celula) -> str | None:
    if celula is None or celula == "":
        return None
    if isinstance(celula, datetime):
        return celula.date().isoformat()
    if isinstance(celula, date):
        return celula.isoformat()
    return None


def _texto(celula) -> str | None:
    if celula is None:
        return None
    texto = str(celula).strip()
    return texto or None


def _inteiro(celula, default: int = 0) -> int:
    try:
        return int(celula)
    except (TypeError, ValueError):
        return default


def _inteiro_ou_none(celula) -> int | None:
    try:
        return int(celula)
    except (TypeError, ValueError):
        return None


def importar_prestadores(wb: openpyxl.Workbook, conn: sqlite3.Connection) -> tuple[int, int]:
    ws = wb["PROJ_PREST"]
    agora = datetime.now().isoformat()
    importados = 0
    ignorados = 0
    for r in range(6, ws.max_row + 1):
        prestador = _texto(ws.cell(row=r, column=4).value)  # D
        data_solicitacao = _valor_data(ws.cell(row=r, column=11).value)  # K
        if not prestador or not data_solicitacao:
            ignorados += 1
            continue

        conn.execute(
            """
            INSERT INTO prestadores (
                item, codigo, prestador, disciplina, disciplina_sla, peps,
                obra_referencia, revisao, num_documentos, data_solicitacao,
                data_limite, data_analise, hold_inicio, hold_fim, num_at,
                revisao_at, responsavel, status_analise, observacoes,
                natureza_revisao, num_erros, etg,
                criado_em, criado_por, atualizado_em, atualizado_por
            ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            """,
            (
                _inteiro_ou_none(ws.cell(row=r, column=2).value),  # B item
                _texto(ws.cell(row=r, column=3).value),  # C codigo
                prestador,
                _texto(ws.cell(row=r, column=5).value),  # E disciplina
                _texto(ws.cell(row=r, column=6).value),  # F disciplina_sla
                _texto(ws.cell(row=r, column=7).value),  # G peps
                _texto(ws.cell(row=r, column=8).value),  # H obra_referencia
                _inteiro(ws.cell(row=r, column=9).value, 0),  # I revisao
                _inteiro(ws.cell(row=r, column=10).value, 0),  # J num_documentos
                data_solicitacao,
                _valor_data(ws.cell(row=r, column=12).value),  # L data_limite
                _valor_data(ws.cell(row=r, column=13).value),  # M data_analise
                _valor_data(ws.cell(row=r, column=14).value),  # N hold_inicio
                _valor_data(ws.cell(row=r, column=15).value),  # O hold_fim
                _texto(ws.cell(row=r, column=19).value),  # S num_at
                _inteiro_ou_none(ws.cell(row=r, column=20).value),  # T revisao_at
                _texto(ws.cell(row=r, column=21).value),  # U responsavel
                _texto(ws.cell(row=r, column=22).value) or "EM ANÁLISE",  # V status_analise
                _texto(ws.cell(row=r, column=23).value),  # W observacoes
                _texto(ws.cell(row=r, column=27).value),  # AA natureza_revisao
                _inteiro_ou_none(ws.cell(row=r, column=28).value),  # AB num_erros
                (_texto(ws.cell(row=r, column=29).value) or "não").upper(),  # AC etg
                agora, USUARIO_IMPORTACAO, agora, USUARIO_IMPORTACAO,
            ),
        )
        importados += 1
    return importados, ignorados


def importar_cessionarios(wb: openpyxl.Workbook, conn: sqlite3.Connection) -> tuple[int, int]:
    ws = wb["PROJ_CESS"]
    agora = datetime.now().isoformat()
    importados = 0
    ignorados = 0
    for r in range(6, ws.max_row + 1):
        cessionario = _texto(ws.cell(row=r, column=4).value)  # D
        data_solicitacao = _valor_data(ws.cell(row=r, column=9).value)  # I
        if not cessionario or not data_solicitacao:
            ignorados += 1
            continue

        conn.execute(
            """
            INSERT INTO cessionarios (
                item, codigo, cessionario, disciplina, disciplina_sla,
                revisao, num_documentos, data_solicitacao, tipo, sla_dias,
                data_limite, data_analise, hold_inicio, hold_fim, num_at,
                revisao_at, responsavel, status_analise, observacoes,
                natureza_revisao, num_erros, etg,
                criado_em, criado_por, atualizado_em, atualizado_por
            ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            """,
            (
                _inteiro_ou_none(ws.cell(row=r, column=2).value),  # B item
                _texto(ws.cell(row=r, column=3).value),  # C codigo
                cessionario,
                _texto(ws.cell(row=r, column=5).value),  # E disciplina
                _texto(ws.cell(row=r, column=6).value),  # F disciplina_sla
                _inteiro(ws.cell(row=r, column=7).value, 0),  # G revisao
                _inteiro(ws.cell(row=r, column=8).value, 0),  # H num_documentos
                data_solicitacao,
                _texto(ws.cell(row=r, column=23).value),  # W tipo
                _inteiro_ou_none(ws.cell(row=r, column=10).value),  # J sla_dias
                _valor_data(ws.cell(row=r, column=11).value),  # K data_limite
                _valor_data(ws.cell(row=r, column=12).value),  # L data_analise
                _valor_data(ws.cell(row=r, column=13).value),  # M hold_inicio
                _valor_data(ws.cell(row=r, column=14).value),  # N hold_fim
                _texto(ws.cell(row=r, column=18).value),  # R num_at
                _inteiro_ou_none(ws.cell(row=r, column=19).value),  # S revisao_at
                _texto(ws.cell(row=r, column=20).value),  # T responsavel
                _texto(ws.cell(row=r, column=21).value) or "EM ANÁLISE",  # U status_analise
                _texto(ws.cell(row=r, column=22).value),  # V observacoes
                _texto(ws.cell(row=r, column=26).value),  # Z natureza_revisao
                _inteiro_ou_none(ws.cell(row=r, column=27).value),  # AA num_erros
                (_texto(ws.cell(row=r, column=28).value) or "não").upper(),  # AB etg
                agora, USUARIO_IMPORTACAO, agora, USUARIO_IMPORTACAO,
            ),
        )
        importados += 1
    return importados, ignorados


def importar_avaliacoes(wb_avaliacao: openpyxl.Workbook | None, conn: sqlite3.Connection) -> tuple[int, int]:
    if wb_avaliacao is None:
        return 0, 0
    ws = wb_avaliacao["AVALIAÇÃO_PRESTADORES"]
    agora = datetime.now().isoformat()
    importados = 0
    ignorados = 0
    for r in range(5, ws.max_row + 1):
        nome_prestador = _texto(ws.cell(row=r, column=2).value)  # B
        data_avaliacao = _valor_data(ws.cell(row=r, column=3).value)  # C
        nota = _inteiro_ou_none(ws.cell(row=r, column=6).value)  # F
        if not nome_prestador or not data_avaliacao or nota is None:
            ignorados += 1
            continue
        conn.execute(
            """
            INSERT INTO avaliacoes_prestadores (
                codigo_prestador, nome_prestador, data_avaliacao, nome_projeto,
                at_referencia, nota, analista_responsavel, observacoes,
                criado_em, criado_por
            ) VALUES (?,?,?,?,?,?,?,?,?,?)
            """,
            (
                _texto(ws.cell(row=r, column=1).value),  # A codigo_prestador
                nome_prestador,
                data_avaliacao,
                _texto(ws.cell(row=r, column=4).value),  # D nome_projeto
                _texto(ws.cell(row=r, column=5).value),  # E at_referencia
                nota,
                _texto(ws.cell(row=r, column=8).value),  # H analista_responsavel
                _texto(ws.cell(row=r, column=9).value),  # I observacoes
                agora, USUARIO_IMPORTACAO,
            ),
        )
        importados += 1
    return importados, ignorados


def main() -> None:
    parser = argparse.ArgumentParser(description="Importa dados históricos da planilha GAT para o SQLite.")
    parser.add_argument("--xlsm", required=True, help="Caminho do arquivo Controle_GAT_Projetos_2026.xlsm")
    parser.add_argument("--avaliacao", required=False, help="Caminho do arquivo Avaliacao_Prestadores_GAT.xlsx")
    parser.add_argument("--destino", required=True, help="Caminho do banco SQLite de destino")
    parser.add_argument("--forcar", action="store_true", help="Sobrescreve mesmo se já houver dados no destino")
    parser.add_argument(
        "--atualizar", action="store_true",
        help="Atualiza os dados a partir de uma planilha mais recente: cria um backup do banco atual, apaga "
             "os registros existentes de prestadores/cessionarios (não mexe em usuários, reuniões, avaliações "
             "de checklist, planos de ação nem configurações) e reimporta do zero a partir da planilha informada.",
    )
    args = parser.parse_args()

    destino = Path(args.destino)

    # Garante que o esquema (tabelas) exista no destino antes de importar.
    # A ordem de import importa: `gat.config.DB_PATH` precisa ser sobrescrito
    # ANTES do primeiro `import gat.database`, pois este faz
    # `from gat.config import DB_PATH` (vínculo por valor, não por referência).
    sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
    import gat.config as config
    config.DB_PATH = destino
    import gat.database as database
    database.init_db()

    conn = sqlite3.connect(destino)
    conn.row_factory = sqlite3.Row

    existentes_prest = conn.execute("SELECT COUNT(*) AS n FROM prestadores").fetchone()["n"]
    existentes_cess = conn.execute("SELECT COUNT(*) AS n FROM cessionarios").fetchone()["n"]
    existentes = existentes_prest + existentes_cess

    if args.atualizar:
        caminho_backup = database.criar_backup()
        if caminho_backup is None and destino.exists():
            print("Não foi possível criar um backup de segurança antes de atualizar — operação cancelada.")
            conn.close()
            sys.exit(1)
        if caminho_backup is not None:
            print(f"Backup de segurança criado em: {caminho_backup}")
        conn.execute("DELETE FROM prestadores")
        conn.execute("DELETE FROM cessionarios")
        conn.commit()
        print(f"Registros anteriores apagados: {existentes_prest} prestador(es), {existentes_cess} cessionário(s).")
    elif existentes > 0 and not args.forcar:
        print(f"O destino já possui {existentes} registro(s) nas tabelas de projetos. "
              f"Use --atualizar para substituir pelos dados de uma planilha mais nova (com backup automático), "
              f"ou --forcar para importar mesmo assim (pode gerar duplicidade).")
        conn.close()
        sys.exit(1)

    print(f"Lendo planilha principal: {args.xlsm}")
    wb = openpyxl.load_workbook(args.xlsm, data_only=True, keep_vba=True)

    wb_avaliacao = None
    if args.avaliacao:
        print(f"Lendo planilha de avaliação: {args.avaliacao}")
        wb_avaliacao = openpyxl.load_workbook(args.avaliacao, data_only=True)

    prest_ok, prest_skip = importar_prestadores(wb, conn)
    cess_ok, cess_skip = importar_cessionarios(wb, conn)
    aval_ok, aval_skip = importar_avaliacoes(wb_avaliacao, conn)

    agora = datetime.now().isoformat()
    evento = "ATUALIZACAO_PLANILHA" if args.atualizar else "IMPORTACAO_INICIAL"
    for tabela, ok in (("prestadores", prest_ok), ("cessionarios", cess_ok), ("avaliacoes_prestadores", aval_ok)):
        if ok:
            conn.execute(
                "INSERT INTO historico_edicoes (tabela, registro_id, campo, valor_anterior, valor_novo, usuario, data_hora) "
                "VALUES (?, 0, ?, NULL, ?, ?, ?)",
                (tabela, evento, f"{ok} registro(s) importado(s) da planilha", USUARIO_IMPORTACAO, agora),
            )

    conn.commit()
    conn.close()

    print("Importação concluída:")
    print(f"  Prestadores:  {prest_ok} importado(s), {prest_skip} linha(s) em branco ignorada(s)")
    print(f"  Cessionários: {cess_ok} importado(s), {cess_skip} linha(s) em branco ignorada(s)")
    print(f"  Avaliações:   {aval_ok} importado(s), {aval_skip} linha(s) em branco ignorada(s)")


if __name__ == "__main__":
    main()
